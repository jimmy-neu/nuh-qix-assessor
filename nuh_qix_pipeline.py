import argparse
import json
import os
import platform
import re
import time
from typing import List, Optional

from dotenv import load_dotenv
from google import genai
from google.genai import types
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# -----------------------------
# Configuration & Setup
# -----------------------------
load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
MODEL_NAME = "gemini-2.5-flash"

FULL_RUBRIC = """
Assessment Score (Minimum): Outstanding-85; Merit-70; Recognition - 50

1. Header (Max Score: 5)
- Meet Expectations (3): Are appropriate team members, leader & sponsor identified to participate?
- Above Expectations (5): Does the title clearly explain the goal of the project?

2. Background (Max Score: 10)
- Meet Expectations (6): Is the nature of the problem clearly explained? Is the problem worthwhile resolving? Is the scope clearly defined and manageable for the Improvement Sprints /project length?
- Above Expectations (10): Is appropriate data collected and presented to quantify and understand the problem?

3. Goal (Max Score: 5)
- Meet Expectations (3): Is the goal SMART? (Specific, measurable, achievable, realistic & time based)
- Above Expectations (5): Is the goal linked to NUH / department objectives and to the problem defined?

4. Problem Analysis (Max Score: 20)
- Meet Expectations (12): Are the appropriate lean / PDCA tools applied to analyze the problems? Does the analysis link back to the problem and goals?
- Above Expectations (20): Are the root causes and main delays/bottlenecks of the problem well identified?

5. Implementation Plan (Max Score: 20)
- Meet Expectations (12): Do the solutions and implementation plan directly address the root causes of problems? Is the implementation plan clear and timely? Does it include specific what, who & when?
- Above Expectations (20): For Improvement Sprints, was a rapid experiment carried out and were insights gained from it? Were some changes started immediately during the event or the next week?

6. Benefits / Results (Max Score: 30)
- Meet Expectations (20): Is there a clear significant improvement? Do the results match the goal? Are the benefits quantified? Are there intangible results? Are clear charts used to show results/improvement? Are dips explained?
- Above Expectations (30): Are the results sustained over the time? (Ideally show results >= 3 months)

7. Follow-up, Spread & Insights (Max Score: 10)
- Meet Expectations (6): Has the team created a plan to overcome any outstanding or potential issues? Has the team created and/or implemented an effective plan to spread the project?
- Above Expectations (10): Has reflection taken place and insights / lessons learned been identified?
"""


# -----------------------------
# Utilities
# -----------------------------
def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()


def autosize_columns(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            cell_len = len(str(cell.value))
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def apply_table_formatting(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize_columns(ws)


# -----------------------------
# PPTX -> PDF Conversion
# -----------------------------
def convert_pptx_folder_to_pdf(input_folder: str, output_folder: str, force: bool = False) -> int:
    input_folder = os.path.abspath(input_folder)
    output_folder = os.path.abspath(output_folder)

    if not os.path.exists(input_folder):
        print(f"Error: Input folder not found: {input_folder}")
        return 0

    pptx_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pptx")]
    if not pptx_files:
        print(f"No .pptx files found in {input_folder}")
        return 0

    if platform.system() != "Windows":
        print("PPTX to PDF conversion requires Windows with PowerPoint installed. Skipping conversion.")
        return 0

    try:
        import comtypes.client
    except Exception as exc:
        print(f"Could not import comtypes. Skipping conversion. Error: {exc}")
        return 0

    ensure_dir(output_folder)

    powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
    powerpoint.Visible = 1

    converted = 0
    try:
        print(f"Scanning '{input_folder}' for .pptx files...")
        for filename in pptx_files:
            input_path = os.path.join(input_folder, filename)
            output_filename = os.path.splitext(filename)[0] + ".pdf"
            output_path = os.path.join(output_folder, output_filename)

            if not force and os.path.exists(output_path):
                print(f"Skipping (already exists): {output_filename}")
                continue

            print(f"Converting: {filename} -> {output_filename}")
            deck = powerpoint.Presentations.Open(input_path)
            deck.SaveAs(output_path, 32)  # 32 is PDF format
            deck.Close()
            converted += 1
    except Exception as exc:
        print(f"An error occurred during conversion: {exc}")
    finally:
        powerpoint.Quit()
        print(f"Done. Converted {converted} files.")

    return converted


# -----------------------------
# Extraction Agent
# -----------------------------
class ProjectExtraction(BaseModel):
    project_title: str = Field(description="The title of the quality improvement project.")
    department: str = Field(description="The hospital department submitting the project.")
    category: str = Field(description="The framework category: 6S, Process Excellence, or Service Experience.")
    problem_statement: str = Field(description="The core problem or initial state being addressed.")
    smart_goals: str = Field(description="Specific, measurable, achievable, realistic, and time-bound goals.")
    methodology: List[str] = Field(description="List of Lean, Design Thinking, or PDCA tools applied.")
    key_results: str = Field(description="Quantitative and qualitative benefits, financial savings, or sustained improvements extracted.")
    follow_up_plan: str = Field(description="Plans to sustain the results or spread the implementation to other departments.")


def extract_clinical_project(client: genai.Client, local_pdf_path: str) -> ProjectExtraction:
    with open(local_pdf_path, "rb") as file:
        doc_data = file.read()

    prompt = """
    You are the Extraction Agent for a hospital's continuous improvement assessment pipeline.
    Analyze the provided project submission document. Your task is to extract the relevant information 
    to populate the required schema. Pay close attention to both unstructured narrative text and 
    visual elements like charts, graphs, and tables to capture the full scope of the methodologies and results.
    """

    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=[
            types.Part.from_bytes(data=doc_data, mime_type="application/pdf"),
            prompt,
        ],
        config={
            "response_mime_type": "application/json",
            "response_schema": ProjectExtraction,
            "temperature": 0.1,
        },
    )

    return response.parsed


# -----------------------------
# Pre-Screening Agent
# -----------------------------
class ScreeningCheck(BaseModel):
    criterion: str = Field(description="The Level-4 exclusion rule being evaluated.")
    violation_found: bool = Field(description="True if the project meets this exclusion criteria.")
    evidence_found: str = Field(description="Specific evidence from the PDF or JSON justifying the decision.")


class ScreeningResult(BaseModel):
    is_eligible: bool = Field(description="Overall eligibility. True if NO violations are found.")
    primary_violation: str = Field(description="The specific rule that failed, or 'None'.")
    detailed_audit: List[ScreeningCheck]


def run_pre_screening(client: genai.Client, json_path: str, pdf_path: str) -> ScreeningResult:
    with open(json_path, "r", encoding="utf-8") as f:
        extracted_data = f.read()

    with open(pdf_path, "rb") as f:
        doc_data = f.read()

    level_4_rules = """
    Your sole responsibility is to be a strict auditor, flagging rule violations and missing criteria. You are looking only for reasons to classify the project as Level 4.
    Task: Analyze the provided project PDF and identify if it meets ANY of the following strict criteria for a Level 4 classification.
    Level 4 'Rejection' Criteria:
    - A solution is already decided (thus not able to apply improvement methodologies).
    - It is a simple just-do-it or straightforward project with solutions like video/print booklets, pamphlets, leaflets, sending reminders, or just reinforcing current processes.
    - It is an IT project (including Excel or programming) that does not involve any actual process change.
    - There is no measurement or documentation of standard work or results.
    - It is a trial or assessment without a concrete implementation.
    - It involves research, randomized control trials, or studies that have pre-empted an intervention.
    - It is just the fine-tuning of new services during a setting-up phase.
    - It's a service development plan like buying new equipment or starting a new service.
    - It's a Regular operation review or fine-tuning (e.g., PAR level, slot allocations, schedules).
    - It's an RCA (Root Cause Analysis) without meaningful supporting data or for a single incident.
    - It's an implementation of Evidence-based practice without measurement.
    """

    prompt = f"""
    You are the Pre-Screening Agent for a hospital's continuous improvement assessment pipeline.
    Evaluate the provided project (JSON data and full PDF) against the Level-4 Exclusionary Rules.

    RULES:
    {level_4_rules}

    EXTRACTED JSON SUMMARY:
    {extracted_data}

    INSTRUCTIONS:
    If you find any violations, you must state the criterion that was met and provide the specific evidence or quote from the document that proves it.
    """

    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=[
            types.Part.from_bytes(data=doc_data, mime_type="application/pdf"),
            prompt,
        ],
        config={
            "response_mime_type": "application/json",
            "response_schema": ScreeningResult,
            "temperature": 0.0,
        },
    )

    return response.parsed


# -----------------------------
# Grading Agent
# -----------------------------
def upload_pdf_to_gemini(client: genai.Client, pdf_path: str):
    print(f"Uploading {pdf_path} to Gemini...")
    pdf_file = client.files.upload(file=pdf_path, config={"mime_type": "application/pdf"})

    while pdf_file.state.name == "PROCESSING":
        print(".", end="", flush=True)
        time.sleep(2)
        pdf_file = client.files.get(name=pdf_file.name)

    if pdf_file.state.name == "FAILED":
        raise ValueError(f"Failed to process PDF: {pdf_file.name}")

    print("\nPDF Uploaded Successfully!")
    return pdf_file


def call_gemini_agent(client: genai.Client, system_instruction: str, pdf_file, text_prompt: str, require_json: bool = False) -> str:
    config = types.GenerateContentConfig(
        system_instruction=system_instruction,
        temperature=0.2,
        response_mime_type="application/json" if require_json else "text/plain",
    )

    contents = [pdf_file, text_prompt]
    response = client.models.generate_content(model=MODEL_NAME, contents=contents, config=config)
    return response.text


def positive_assessor(client: genai.Client, pdf_file, json_text: str) -> str:
    print("-> Positive Assessor (Defense) analyzing...")
    prompt = f"""You are the Positive Advocate for this clinical project. 
    Using BOTH the provided PDF document and the JSON summary, review the project against this rubric:
    {FULL_RUBRIC}

    Your goal is to highlight all strengths. Find specific evidence, charts, or quotes in the PDF and JSON that justify 'Above Expectations' scores for every category. Formulate a strong defensive argument."""

    text_prompt = f"EXTRACTED PROJECT JSON:\n{json_text}\n\n{prompt}"
    return call_gemini_agent(client, "Act as a strict but highly supportive positive advocate.", pdf_file, text_prompt)


def negative_assessor(client: genai.Client, pdf_file, json_text: str) -> str:
    print("-> Negative Assessor (Prosecution) analyzing...")
    prompt = f"""You are a ruthless, veteran clinical auditor known for extreme strictness. 
    Using BOTH the provided PDF document and the JSON summary, review the project against this rubric:
    {FULL_RUBRIC}

    Your goal is to aggressively drag the score down. You must actively search for technicalities, missing long-term data, or subjective claims. 
    - If a criterion requires multiple elements (e.g., 'quantified benefits' AND 'intangible results'), and they only have one, aggressively attack the missing element.
    - If they claim 'sustained results', strictly verify if the charts in the PDF explicitly prove >= 3 months. If it's only 2.5 months, flag it as a failure.
    - Argue fiercely why this project DOES NOT deserve 'Above Expectations' and must be capped at 'Meet Expectations' or 'Below Expectations'."""

    text_prompt = f"EXTRACTED PROJECT JSON:\n{json_text}\n\n{prompt}"
    return call_gemini_agent(client, "Act as a hostile, highly critical project auditor who penalizes missing data heavily.", pdf_file, text_prompt)


def independent_judge(client: genai.Client, pdf_file, json_text: str, pos_arg: str, neg_arg: str) -> dict:
    print("-> Independent Judge finalizing scores...")
    system_prompt = f"""You are the Lead Meta-Judge for the NUH QIX awards. 
    You have the original PDF submission, the extracted JSON, a Positive Advocate's review, and a Strict Skeptic's review.

    RUBRIC & THRESHOLDS:
    {FULL_RUBRIC}
    - Outstanding: >= 85 (HARD REQUIREMENT: Only the top 10% of elite projects achieve this. Evidence must be flawless.)
    - Merit: >= 70
    - Recognition: >= 50

    CRITICAL GRADING RULES:
    1. THE BURDEN OF PROOF: You MUST default to 'Meet Expectations' or lower. You are strictly forbidden from awarding 'Above Expectations' unless the Positive Advocate provides explicit, undeniable quotes/charts from the source documents that satisfy EVERY SINGLE requirement in that rubric tier.
    2. PENALIZE VAGUENESS: If the Skeptic successfully points out that a claim is subjective, unquantified, or lacks a specific timeframe, you MUST downgrade the score. 
    3. EXACT DISCRETE SCORES: You must only select the exact integer scores provided in the rubric (e.g., for Background, you must choose 3, 6, or 10. Do not invent a score like 8 or 9).

    INSTRUCTIONS:
    Cross-reference the debate. Rule on each category. Output strictly matching this JSON schema:
    {{
      "assessments": [
        {{
          "category": "String (e.g., '1. Background')",
          "max_score": "Integer (e.g., 10)",
          "ai_score": "Integer (Must be an exact discrete score from the rubric)",
          "ai_justification": "String (Explain why you rejected the higher score, or why the evidence was so flawless it forced you to award it)",
          "extracted_quote": "String (Exact quote from the PDF/JSON supporting the score)"
        }}
      ]
    }}"""

    debate_context = f"""
    --- EXTRACTED JSON ---
    {json_text}

    --- POSITIVE ADVOCATE ARGUMENT ---
    {pos_arg}

    --- STRICT SKEPTIC ARGUMENT ---
    {neg_arg}
    """

    raw_json = call_gemini_agent(client, system_prompt, pdf_file, debate_context, require_json=True)
    return json.loads(raw_json)


def grade_project(client: genai.Client, pdf_filepath: str, json_filepath: str) -> dict:
    with open(json_filepath, "r") as f:
        project_json = json.load(f)
        json_text = json.dumps(project_json, indent=2)

    print(f"Starting Multi-Agent Grading for: {project_json.get('project_title', 'Unknown')}")

    pdf_file = upload_pdf_to_gemini(client, pdf_filepath)

    try:
        pos_arg = positive_assessor(client, pdf_file, json_text)
        neg_arg = negative_assessor(client, pdf_file, json_text)
        final_assessment = independent_judge(client, pdf_file, json_text, pos_arg, neg_arg)

        total_score = sum(int(item["ai_score"]) for item in final_assessment["assessments"])

        if total_score >= 85:
            label = "Outstanding"
        elif total_score >= 70:
            label = "Merit"
        elif total_score >= 50:
            label = "Recognition"
        else:
            label = "Below Recognition"

        final_assessment["total_score"] = total_score
        final_assessment["label"] = label

        print(f"\n✅ Assessment Complete! Final Score: {total_score}/100 ({label})")
        return final_assessment
    finally:
        print("Cleaning up PDF from Gemini servers...")
        client.files.delete(name=pdf_file.name)
        print("Cleanup successful.")


# -----------------------------
# Excel Output
# -----------------------------
def write_excel(
    output_path: str,
    extraction_rows: List[List],
    prescreen_summary_rows: List[List],
    prescreen_detail_rows: List[List],
    grading_detail_rows: List[List],
    summary_entries: List[dict],
) -> None:
    wb = Workbook()

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "Project ID",
        "PDF File",
        "Project Title",
        "Status",
        "Eligibility",
        "Level 4 Reason",
        "AI Total Score",
        "AI Label",
        "Final Total Score",
        "Final Label",
        "Error",
    ]
    ws_summary.append(summary_headers)

    for entry in summary_entries:
        ws_summary.append(
            [
                entry.get("project_id"),
                entry.get("pdf_file"),
                entry.get("project_title"),
                entry.get("status"),
                entry.get("eligibility"),
                entry.get("level4_reason"),
                entry.get("ai_total_score"),
                entry.get("ai_label"),
                None,
                None,
                entry.get("error"),
            ]
        )

    for row_idx in range(2, ws_summary.max_row + 1):
        status = ws_summary.cell(row=row_idx, column=4).value
        eligibility = ws_summary.cell(row=row_idx, column=5).value

        if status == "graded":
            ws_summary.cell(
                row=row_idx, column=9
            ).value = f"=SUMIF(Grading_Details!A:A,A{row_idx},Grading_Details!G:G)"
            ws_summary.cell(
                row=row_idx, column=10
            ).value = (
                f'=IF(I{row_idx}>=85,"Outstanding",'
                f'IF(I{row_idx}>=70,"Merit",'
                f'IF(I{row_idx}>=50,"Recognition","Below Recognition")))'
            )
        elif eligibility == "Ineligible":
            ws_summary.cell(row=row_idx, column=10).value = "LEVEL 4"

    apply_table_formatting(ws_summary)

    # Extraction Sheet
    ws_extract = wb.create_sheet("Extraction")
    extract_headers = [
        "Project ID",
        "PDF File",
        "Project Title",
        "Department",
        "Category",
        "Problem Statement",
        "SMART Goals",
        "Methodology",
        "Key Results",
        "Follow Up Plan",
        "Error",
    ]
    ws_extract.append(extract_headers)
    for row in extraction_rows:
        ws_extract.append(row)
    apply_table_formatting(ws_extract)

    # Pre-Screening Summary Sheet
    ws_screen = wb.create_sheet("PreScreening")
    prescreen_headers = [
        "Project ID",
        "PDF File",
        "Eligible",
        "Primary Violation",
        "Violation Evidence",
        "Error",
    ]
    ws_screen.append(prescreen_headers)
    for row in prescreen_summary_rows:
        ws_screen.append(row)
    apply_table_formatting(ws_screen)

    # Pre-Screening Details Sheet
    ws_screen_detail = wb.create_sheet("PreScreening_Details")
    prescreen_detail_headers = [
        "Project ID",
        "Criterion",
        "Violation Found",
        "Evidence",
    ]
    ws_screen_detail.append(prescreen_detail_headers)
    for row in prescreen_detail_rows:
        ws_screen_detail.append(row)
    apply_table_formatting(ws_screen_detail)

    # Grading Details Sheet
    ws_grade_detail = wb.create_sheet("Grading_Details")
    grading_headers = [
        "Project ID",
        "PDF File",
        "Category",
        "Max Score",
        "AI Score",
        "Human Score",
        "Final Score",
        "AI Justification",
        "Extracted Quote",
    ]
    ws_grade_detail.append(grading_headers)

    for row in grading_detail_rows:
        ws_grade_detail.append(row)

    for row_idx in range(2, ws_grade_detail.max_row + 1):
        ws_grade_detail.cell(
            row=row_idx, column=7
        ).value = f"=IF(F{row_idx}=\"\",E{row_idx},F{row_idx})"

    apply_table_formatting(ws_grade_detail)

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


# -----------------------------
# Main Pipeline
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="NUH-QIX end-to-end assessment pipeline.")
    parser.add_argument("--pptx_dir", default="./project_pptx", help="Folder containing .pptx files")
    parser.add_argument("--pdf_dir", default="./project", help="Folder to store converted PDFs")
    parser.add_argument("--extract_dir", default="./extracted_results", help="Folder to store extracted JSON files")
    parser.add_argument("--output_excel", default="./assessment_results.xlsx", help="Excel report output path")
    parser.add_argument("--skip_pptx", action="store_true", help="Skip PPTX to PDF conversion")
    parser.add_argument("--force_convert", action="store_true", help="Force reconversion of PPTX files")
    args = parser.parse_args()

    if not GEMINI_API_KEY:
        print("Error: GEMINI_API_KEY is not set. Please set it in your environment or .env file.")
        return

    ensure_dir(args.extract_dir)

    if not args.skip_pptx:
        convert_pptx_folder_to_pdf(args.pptx_dir, args.pdf_dir, force=args.force_convert)

    # Build list of PDFs to process based on PPTX names
    pptx_files = [f for f in os.listdir(args.pptx_dir) if f.lower().endswith(".pptx")] if os.path.exists(args.pptx_dir) else []
    pptx_bases = {os.path.splitext(f)[0] for f in pptx_files}

    pdf_files: List[str] = []
    if pptx_bases:
        for base in sorted(pptx_bases):
            pdf_path = os.path.join(args.pdf_dir, f"{base}.pdf")
            if os.path.exists(pdf_path):
                pdf_files.append(pdf_path)
            else:
                print(f"Warning: Missing PDF for {base}.pptx in {args.pdf_dir}")
    else:
        # Fallback: process any PDFs in the pdf_dir if no PPTX files were found
        if os.path.exists(args.pdf_dir):
            pdf_files = [
                os.path.join(args.pdf_dir, f)
                for f in os.listdir(args.pdf_dir)
                if f.lower().endswith(".pdf")
            ]

    if not pdf_files:
        print("No PDFs found to process. Ensure PPTX conversion succeeded or place PDFs in the pdf directory.")
        return

    client = genai.Client(api_key=GEMINI_API_KEY)

    extraction_rows: List[List] = []
    prescreen_summary_rows: List[List] = []
    prescreen_detail_rows: List[List] = []
    grading_detail_rows: List[List] = []
    summary_entries: List[dict] = []

    for pdf_path in pdf_files:
        pdf_file = os.path.basename(pdf_path)
        project_id = sanitize_filename(os.path.splitext(pdf_file)[0])
        project_title = project_id

        status = ""
        eligibility = ""
        level4_reason = ""
        ai_total_score = ""
        ai_label = ""
        error = ""

        extraction_data: Optional[ProjectExtraction] = None
        json_path = os.path.join(args.extract_dir, f"{project_id}.json")

        try:
            extraction_data = extract_clinical_project(client, pdf_path)
            project_title = extraction_data.project_title or project_title
            with open(json_path, "w", encoding="utf-8") as json_file:
                json_file.write(extraction_data.model_dump_json(indent=2))
        except Exception as exc:
            status = "extraction_failed"
            error = str(exc)

        extraction_rows.append(
            [
                project_id,
                pdf_file,
                getattr(extraction_data, "project_title", ""),
                getattr(extraction_data, "department", ""),
                getattr(extraction_data, "category", ""),
                getattr(extraction_data, "problem_statement", ""),
                getattr(extraction_data, "smart_goals", ""),
                "; ".join(getattr(extraction_data, "methodology", []) or []),
                getattr(extraction_data, "key_results", ""),
                getattr(extraction_data, "follow_up_plan", ""),
                error if status == "extraction_failed" else "",
            ]
        )

        if not extraction_data:
            summary_entries.append(
                {
                    "project_id": project_id,
                    "pdf_file": pdf_file,
                    "project_title": project_title,
                    "status": status,
                    "eligibility": eligibility,
                    "level4_reason": level4_reason,
                    "ai_total_score": ai_total_score,
                    "ai_label": ai_label,
                    "error": error,
                }
            )
            continue

        try:
            screening = run_pre_screening(client, json_path, pdf_path)
            eligibility = "Eligible" if screening.is_eligible else "Ineligible"
            level4_reason = screening.primary_violation if not screening.is_eligible else ""

            violations = [
                f"{check.criterion}: {check.evidence_found}"
                for check in screening.detailed_audit
                if check.violation_found
            ]
            prescreen_summary_rows.append(
                [
                    project_id,
                    pdf_file,
                    eligibility,
                    screening.primary_violation,
                    " | ".join(violations),
                    "",
                ]
            )

            for check in screening.detailed_audit:
                prescreen_detail_rows.append(
                    [
                        project_id,
                        check.criterion,
                        check.violation_found,
                        check.evidence_found,
                    ]
                )
        except Exception as exc:
            status = "screening_failed"
            error = str(exc)
            prescreen_summary_rows.append([project_id, pdf_file, "", "", "", error])
            summary_entries.append(
                {
                    "project_id": project_id,
                    "pdf_file": pdf_file,
                    "project_title": project_title,
                    "status": status,
                    "eligibility": eligibility,
                    "level4_reason": level4_reason,
                    "ai_total_score": ai_total_score,
                    "ai_label": ai_label,
                    "error": error,
                }
            )
            continue

        if eligibility == "Ineligible":
            status = "level4_ineligible"
            summary_entries.append(
                {
                    "project_id": project_id,
                    "pdf_file": pdf_file,
                    "project_title": project_title,
                    "status": status,
                    "eligibility": eligibility,
                    "level4_reason": level4_reason,
                    "ai_total_score": ai_total_score,
                    "ai_label": ai_label,
                    "error": "",
                }
            )
            continue

        try:
            grading = grade_project(client, pdf_path, json_path)
            ai_total_score = grading.get("total_score", "")
            ai_label = grading.get("label", "")

            for item in grading.get("assessments", []):
                grading_detail_rows.append(
                    [
                        project_id,
                        pdf_file,
                        item.get("category"),
                        item.get("max_score"),
                        item.get("ai_score"),
                        "",
                        "",
                        item.get("ai_justification"),
                        item.get("extracted_quote"),
                    ]
                )

            status = "graded"
        except Exception as exc:
            status = "grading_failed"
            error = str(exc)

        summary_entries.append(
            {
                "project_id": project_id,
                "pdf_file": pdf_file,
                "project_title": project_title,
                "status": status,
                "eligibility": eligibility,
                "level4_reason": level4_reason,
                "ai_total_score": ai_total_score,
                "ai_label": ai_label,
                "error": error,
            }
        )

    write_excel(
        args.output_excel,
        extraction_rows,
        prescreen_summary_rows,
        prescreen_detail_rows,
        grading_detail_rows,
        summary_entries,
    )


if __name__ == "__main__":
    main()
